#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para extrair texto de PDF e exibir linha por linha na tela
"""

import sys
import os
import re
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import csv

try:
    import PyPDF2
except ImportError:
    print("PyPDF2 não encontrado. Instalando...")
    os.system("pip install PyPDF2")
    import PyPDF2

def extrair_texto_pdf(caminho_pdf, debug=False):
    """
    Extrai texto de um arquivo PDF
    
    Args:
        caminho_pdf (str): Caminho para o arquivo PDF
        debug (bool): Se True, exibe informações detalhadas
        
    Returns:
        str: Texto extraído do PDF
    """
    try:
        with open(caminho_pdf, 'rb') as arquivo:
            leitor_pdf = PyPDF2.PdfReader(arquivo)
            texto_completo = ""
            
            if debug:
                print(f"PDF possui {len(leitor_pdf.pages)} páginas")
            
            for numero_pagina, pagina in enumerate(leitor_pdf.pages, 1):
                if debug:
                    print(f"\n--- PÁGINA {numero_pagina} ---")
                texto_pagina = pagina.extract_text()
                texto_pagina = texto_pagina.replace('\n', '|')
                texto_completo += texto_pagina
                
                if debug:
                    # Exibir linha por linha
                    linhas = texto_pagina.split('|')
                    for i, linha in enumerate(linhas, 1):
                        if linha.strip():  # Só exibe linhas não vazias
                            print(f"Linha {i}: {linha}")
                
            return texto_completo
            
    except FileNotFoundError:
        print(f"Erro: Arquivo '{caminho_pdf}' não encontrado.")
        return None
    except Exception as e:
        print(f"Erro ao processar o PDF: {e}")
        return None

def organizar_lancamentos_por_data(texto):
    """
    Organiza os lançamentos em uma estrutura de dados baseada na data (dd/mm)
    
    Args:
        texto (str): Texto extraído do PDF
        
    Returns:
        dict: Dicionário com lançamentos organizados por data
    """
    # Padrão para identificar datas no formato dd/mm
    padrao_data = r'^\d{2}/\d{2}'
    
    # Dividir o texto em linhas e limpar
    linhas = texto.split('|')
    linhas_limpas = []
    
    # Limpar e processar linhas
    for linha in linhas:
        linha = linha.strip()
        if not linha:  # Pular linhas vazias
            continue
        
        # Verificar se a linha contém múltiplas datas (problema de concatenação)
        datas_encontradas = re.findall(r'\d{2}/\d{2}', linha)
        if len(datas_encontradas) > 1:
            # Dividir a linha em múltiplas partes baseado nas datas
            partes = re.split(r'(\d{2}/\d{2})', linha)
            for i in range(1, len(partes), 2):  # Pular o primeiro elemento vazio
                if i + 1 < len(partes):
                    nova_linha = partes[i] + partes[i + 1]
                    if nova_linha.strip():
                        linhas_limpas.append(nova_linha.strip())
        else:
            # Verificar se há uma data no meio da linha (problema de concatenação)
            match_data_meio = re.search(r'(\d{2}/\d{2})', linha)
            if match_data_meio and not linha.startswith(match_data_meio.group(1)):
                # Encontrar a posição da data no meio da linha
                pos_data = linha.find(match_data_meio.group(1))
                if pos_data > 0:
                    # Separar em duas linhas
                    linha_antes = linha[:pos_data].strip()
                    linha_depois = linha[pos_data:].strip()
                    if linha_antes:
                        linhas_limpas.append(linha_antes)
                    if linha_depois:
                        linhas_limpas.append(linha_depois)
                else:
                    linhas_limpas.append(linha)
            else:
                linhas_limpas.append(linha)
    
    # Dicionário para armazenar os lançamentos por data
    lancamentos_por_data = {}
    lancamento_atual = []
    data_atual = None
    
    for linha in linhas_limpas:
        # Verificar se a linha começa com uma data (dd/mm)
        if re.match(padrao_data, linha):
            # Se já temos um lançamento em andamento, salvá-lo
            if data_atual and lancamento_atual:
                if data_atual not in lancamentos_por_data:
                    lancamentos_por_data[data_atual] = []
                # Juntar as linhas do lançamento em uma única string
                lancamento_completo = ' | '.join(lancamento_atual)
                lancamentos_por_data[data_atual].append(lancamento_completo)
            
            # Iniciar novo lançamento
            data_atual = linha[:5]  # Pegar apenas dd/mm
            lancamento_atual = [linha]
        else:
            # Adicionar linha ao lançamento atual
            if data_atual:
                lancamento_atual.append(linha)
    
    # Adicionar o último lançamento
    if data_atual and lancamento_atual:
        if data_atual not in lancamentos_por_data:
            lancamentos_por_data[data_atual] = []
        # Juntar as linhas do lançamento em uma única string
        lancamento_completo = ' '.join(lancamento_atual)
        lancamentos_por_data[data_atual].append(lancamento_completo)
    
    return lancamentos_por_data

def exibir_lancamentos_organizados(lancamentos_por_data, debug=False):
    """
    Exibe os lançamentos organizados por data
    
    Args:
        lancamentos_por_data (dict): Dicionário com lançamentos organizados
        debug (bool): Se True, exibe a lista de lançamentos
    """
    if debug:
        print("\n" + "="*60)
        print("LISTA DE LANÇAMENTOS")
        print("="*60)
        
        # Ordenar as datas
        datas_ordenadas = sorted(lancamentos_por_data.keys())
        
        for data in datas_ordenadas:
            lancamentos = lancamentos_por_data[data]
            for lancamento in lancamentos:
                print(lancamento)
        
        print(f"\n📊 Total de datas: {len(datas_ordenadas)}")
        total_lancamentos = sum(len(lancamentos) for lancamentos in lancamentos_por_data.values())
        print(f"📊 Total de lançamentos: {total_lancamentos}")

def extrair_data_valor(lancamento):
    """
    Extrai data, valor e tipo de um lançamento no novo formato
    Valor negativo para D, positivo para C.
    Args:
        lancamento (str): String do lançamento
    Returns:
        tuple: (data_processada, valor_processado, tipo_operacao)
    """
    import re
    
    # Dividir a linha pelos separadores |
    partes = lancamento.split('|')
    
    # Primeira parte contém data e valor
    primeira_parte = partes[0].strip()

    if len(partes) > 1:
        segunda_parte = partes[1].strip()
    else:
        segunda_parte = ""
    
    # Padrão para data (dd/mm)
    padrao_data = r'^(\d{2}/\d{2})'
    
    # Extrair data
    match_data = re.search(padrao_data, primeira_parte)
    if match_data:
        data_dd_mm = match_data.group(1)
        data_processada = f"{data_dd_mm}/2025"
    else:
        data_processada = None
    

    # Padrão para valor monetário (com D ou C concatenado ou separado)
    padrao_valor = r'(\d{1,3}(?:\.\d{3})*,\d{2})(?:\s*\|\s*([DC])|([DC]))'
    
    # Procurar por valores monetários na primeira parte
    match_valor = re.search(padrao_valor, primeira_parte + "|" + segunda_parte)

    valor_processado = None
    tipo_operacao = None
    
    if match_valor:
        valor_str = match_valor.group(1)
        tipo_concat = match_valor.group(2) or match_valor.group(3) or None
        valor_limpo = valor_str.replace('.', '').replace(',', '.')
        try:
            valor_float = float(valor_limpo)
            
            # Buscar tipo: concatenado ou na segunda parte
            tipo = tipo_concat
            if not tipo and len(partes) > 1:
                # Procurar D ou C na segunda parte
                segunda_parte = partes[1].strip()
                if segunda_parte in ['D', 'C']:
                    tipo = segunda_parte
            
            if tipo == 'D':
                valor_float = -valor_float
            elif tipo == 'C':
                valor_float = abs(valor_float)
            else:
                tipo = None
                
            valor_processado = valor_float
            tipo_operacao = tipo
        except ValueError:
            pass
    
    return data_processada, valor_processado, tipo_operacao

def extrair_cnpj_cpf(texto):
    """
    Extrai o primeiro CNPJ ou CPF encontrado em um texto.
    Retorna string vazia se não encontrar.
    Aceita CNPJ com barra ou espaço e CPF com asteriscos.
    Busca apenas a partir da segunda parte da linha (ignora data/valor).
    """
    import re
    
    # Dividir a linha pelos separadores | para buscar em cada parte
    partes = texto.split('|')
    
    # Padrão CNPJ: 00.000.000/0000-00 ou 00.000.000 0000-00
    padrao_cnpj = r'\d{2}\.\d{3}\.\d{3}[ /]\d{4}-\d{2}'
    # Padrão CPF: 000.000.000-00 ou ***.000.000-00 ou ***.000.000-000 ou ***.000.000-**
    padrao_cpf = r'(?:\*{3}|\d{3})\.\d{3}\.\d{3}-(?:\d{2,3}|\*{2})'
    
    # Buscar a partir da segunda parte (ignorar primeira parte com data/valor)
    for i in range(1, len(partes)):
        parte = partes[i].strip()
        # Buscar CNPJ primeiro
        match_cnpj = re.search(padrao_cnpj, parte)
        if match_cnpj:
            return match_cnpj.group(0)
        # Buscar CPF
        match_cpf = re.search(padrao_cpf, parte)
        if match_cpf:
            return match_cpf.group(0)
    
    return ''

def extrair_saldo_final_resumo(lancamento):
    """
    Extrai o saldo final do resumo que aparece no formato "SALDO EM C.CORRENTE(+): 9.185,57C"
    """
    import re
    
    # Padrão para saldo em conta corrente
    padrao_saldo = r'SALDO EM C\.CORRENTE\(\+\):\s*(\d{1,3}(?:\.\d{3})*,\d{2})'
    match = re.search(padrao_saldo, lancamento)
    if match:
        valor_str = match.group(1)
        valor_limpo = valor_str.replace('.', '').replace(',', '.')
        try:
            return float(valor_limpo)
        except ValueError:
            pass
    
    # Padrão para saldo disponível
    padrao_disponivel = r'SALDO DISPONÍVEL\(=\):\s*(\d{1,3}(?:\.\d{3})*,\d{2})'
    match = re.search(padrao_disponivel, lancamento)
    if match:
        valor_str = match.group(1)
        valor_limpo = valor_str.replace('.', '').replace(',', '.')
        try:
            return float(valor_limpo)
        except ValueError:
            pass
    
    return None

def processar_lancamentos_com_data_valor(lancamentos_por_data):
    """
    Processa os lançamentos extraindo data, valor, tipo, cnpj/cpf, documento e pagador/recebedor
    """
    lancamentos_processados = []
    saldo_anterior = None
    ultimo_saldo_dia = None
    
    # Filtrar datas válidas antes de ordenar
    datas_validas = []
    for data in lancamentos_por_data.keys():
        try:
            # Verificar se a data é válida
            dia, mes = data.split('/')
            if 1 <= int(dia) <= 31 and 1 <= int(mes) <= 12:
                datas_validas.append(data)
        except (ValueError, IndexError):
            # Pular datas inválidas
            continue
    
    datas_ordenadas = sorted(datas_validas, 
                           key=lambda x: datetime.strptime(x, '%d/%m'))
    
    for data in datas_ordenadas:
        lancamentos = lancamentos_por_data[data]
        for lancamento in lancamentos:
            data_processada, valor_processado, tipo_operacao = extrair_data_valor(lancamento)
            cnpj_cpf = extrair_cnpj_cpf(lancamento)
            documento = extrair_documento(lancamento)
            pagador_recebedor = extrair_pagador_recebedor(lancamento, cnpj_cpf)
            print(f"Data: {data_processada}, Valor: {valor_processado}, Tipo: {tipo_operacao}")
            
            # Verificar se é saldo anterior
            if "SALDO ANTERIOR" in lancamento and valor_processado is not None:
                saldo_anterior = valor_processado
                continue
            
            # Verificar se é saldo do dia (mas não o último, apenas para referência)
            if "SALDO DO DIA" in lancamento and valor_processado is not None:
                # Não usar este valor como saldo final, apenas para referência
                continue
            
            # Verificar se é o saldo final real (que aparece no resumo)
            if "SALDO EM C.CORRENTE" in lancamento or "SALDO DISPONÍVEL" in lancamento:
                saldo_final = extrair_saldo_final_resumo(lancamento)
                if saldo_final is not None:
                    ultimo_saldo_dia = saldo_final
                    continue
            
            if (data_processada and 
                valor_processado is not None and 
                tipo_operacao is not None and
                valor_processado != 0):
                lancamentos_processados.append({
                    'data': data_processada,
                    'valor': valor_processado,
                    'tipo': tipo_operacao,
                    'cnpj_cpf': cnpj_cpf,
                    'documento': documento,
                    'pagador_recebedor': pagador_recebedor,
                    'lancamento_completo': lancamento
                })
    
    # Se não encontrou o saldo final no resumo, usar o último saldo do dia
    if ultimo_saldo_dia is None:
        for data in reversed(datas_ordenadas):
            lancamentos = lancamentos_por_data[data]
            for lancamento in lancamentos:
                if "SALDO DO DIA" in lancamento:
                    data_processada, valor_processado, tipo_operacao = extrair_data_valor(lancamento)
                    if valor_processado is not None:
                        ultimo_saldo_dia = valor_processado
                        break
            if ultimo_saldo_dia is not None:
                break
    
    return lancamentos_processados, saldo_anterior, ultimo_saldo_dia

def exibir_lancamentos_processados(lancamentos_processados, saldo_anterior, ultimo_saldo_dia, debug=False):
    """
    Exibe os lançamentos processados de forma organizada
    
    Args:
        lancamentos_processados (list): Lista de lançamentos processados
        saldo_anterior (float): Saldo anterior do período
        ultimo_saldo_dia (float): Último saldo do dia
        debug (bool): Se True, exibe a lista detalhada de lançamentos
    """
    print("\n" + "="*80)
    print("CONFERÊNCIA DE SALDOS")
    print("="*80)
    
    # Exibir informações de saldo para conferência
    if saldo_anterior is not None:
        print(f"💰 Saldo Anterior do Período: R$ {saldo_anterior:10.2f}")
    if ultimo_saldo_dia is not None:
        print(f"💰 Saldo Final do Extrato: R$ {ultimo_saldo_dia:10.2f}")
    
    # Calcular saldo das movimentações, entradas e saídas
    if lancamentos_processados:
        saldo_movimentacoes = sum(l['valor'] for l in lancamentos_processados)
        total_entradas = sum(l['valor'] for l in lancamentos_processados if l['valor'] > 0)
        total_saidas = sum(l['valor'] for l in lancamentos_processados if l['valor'] < 0)
        print(f"💳 Saldo das Movimentações: R$ {saldo_movimentacoes:10.2f}")
        print(f"⬆️  Total de Entradas:      R$ {total_entradas:10.2f}")
        print(f"⬇️  Total de Saídas:        R$ {total_saidas:10.2f}")
        
        # Verificar se bate com o saldo do dia
        if saldo_anterior is not None and ultimo_saldo_dia is not None:
            saldo_calculado = saldo_anterior + saldo_movimentacoes
            print(f"🧮 Saldo Calculado (Anterior + Movimentações): R$ {saldo_calculado:10.2f}")
            diferenca = saldo_calculado - ultimo_saldo_dia
            print(f"🔍 Diferença: R$ {diferenca:10.2f}")
            if abs(diferenca) < 0.01:
                print("✅ CONFERÊNCIA: Saldos batem!")
            else:
                print("❌ CONFERÊNCIA: Saldos não batem!")
    print("-" * 80)
    
    # Exibir lista detalhada apenas em modo debug
    if debug:
        print("\n" + "="*80)
        print("LANÇAMENTOS COM DATA E VALOR PROCESSADOS")
        print("="*80)
        
        if not lancamentos_processados:
            print("❌ Nenhum lançamento válido encontrado!")
            return
        
        # Ordenar lançamentos cronologicamente
        lancamentos_ordenados = sorted(lancamentos_processados, key=lambda x: datetime.strptime(x['data'], '%d/%m/%Y'))
        
        print(f"📊 Total de lançamentos: {len(lancamentos_ordenados)}")
        print("-" * 80)
        print(f"{'#':>3}  {'Data':<10}  {'Valor':>12}  {'Tipo':<4}  {'CNPJ/CPF':<20}  {'Documento':<15}  {'Pagador/Recebedor':<20}  Lançamento")
        print("-" * 80)
        for i, lancamento in enumerate(lancamentos_ordenados, 1):
            data = lancamento['data']
            valor = lancamento['valor']
            tipo = lancamento['tipo']
            cnpj_cpf = lancamento.get('cnpj_cpf', '')
            documento = lancamento.get('documento', '')
            pagador_recebedor = lancamento.get('pagador_recebedor', '')
            texto = lancamento['lancamento_completo']
            # Formatar valor com sinal
            sinal = "+" if valor > 0 else ""
            valor_formatado = f"{sinal}{valor:.2f}"
            print(f"{i:3d}. {data:<10}  {valor_formatado:>12}  {tipo:<4}  {cnpj_cpf:<20}  {documento:<15}  {pagador_recebedor:<20}  {texto}")
        print("-" * 80)

def extrair_documento(texto):
    """
    Extrai o documento que aparece após "DOC.:" em um texto.
    Retorna string vazia se não encontrar.
    """
    import re
    
    # Padrão para encontrar "DOC.:" seguido do documento
    padrao = r'DOC\.:\s*([^|]+)'
    
    match = re.search(padrao, texto)
    if match:
        return match.group(1).strip()
    
    return ''

def extrair_pagador_recebedor(texto, cnpj_cpf):
    """
    Extrai o pagador/recebedor do texto.
    Se encontrar "Pagamento Pix", retorna o CNPJ/CPF.
    Caso contrário, busca na segunda, terceira ou quarta posição.
    """
    import re
    
    # Dividir a linha pelos separadores |
    partes = texto.split('|')
    
    # Se for "Pagamento Pix", retorna o CNPJ/CPF
    if 'Pagamento Pix' in texto:
        return cnpj_cpf
    
    # Buscar na segunda, terceira ou quarta posição
    for i in range(1, min(5, len(partes))):
        parte = partes[i].strip()
        
        # Pular se for apenas o tipo (C ou D)
        if parte in ['C', 'D']:
            continue
            
        # Pular se for "Pagamento Pix" ou "Recebimento Pix"
        if parte in ['Pagamento Pix', 'Recebimento Pix']:
            continue
            
        # Pular se for um CNPJ/CPF (já temos essa informação)
        if re.search(r'(?:\*{3}|\d{3})\.\d{3}\.\d{3}-(?:\d{2,3}|\*{2})', parte):
            continue
        if re.search(r'\d{2}\.\d{3}\.\d{3}[ /]\d{4}-\d{2}', parte):
            continue
            
        # Pular se for "DOC.:" ou começar com "DOC.:"
        if parte.startswith('DOC.:'):
            continue
            
        # Se chegou aqui, é provavelmente o pagador/recebedor
        if parte and len(parte) > 2:  # Evitar partes muito pequenas
            return parte
    
    return ''

def gerar_csv_simplificado(lancamentos_processados, caminho_pdf):
    """
    Gera um arquivo CSV (.csv) com apenas as colunas: data, nome, valor
    O nome inclui o sufixo "- SICOOB"
    """
    if not lancamentos_processados:
        print("❌ Nenhum lançamento para gerar CSV!")
        return
    
    # Ordenar lançamentos cronologicamente
    lancamentos_ordenados = sorted(lancamentos_processados, key=lambda x: datetime.strptime(x['data'], '%d/%m/%Y'))
    
    # Obter diretório e nome do arquivo PDF
    pdf_path = Path(caminho_pdf)
    diretorio = pdf_path.parent
    nome_base = pdf_path.stem
    
    # Nome do arquivo CSV
    nome_csv = f"{nome_base}_EXTRATO_SIMPLIFICADO.csv"
    caminho_csv = diretorio / nome_csv
    
    # Criar dados para CSV
    dados_csv = []
    for lancamento in lancamentos_ordenados:
        nome = lancamento.get('pagador_recebedor', '').strip()
        cnpj_cpf = lancamento.get('cnpj_cpf', '').strip()
        documento = lancamento.get('documento', '').strip()
        
        # Se nome e CNPJ/CPF estiverem em branco, usar o documento
        if not nome and not cnpj_cpf and documento:
            nome_com_sufixo = f"{documento} - SICOOB"
        elif nome and cnpj_cpf:
            # Verificar se o nome já contém o CNPJ/CPF para evitar duplicação
            if cnpj_cpf in nome:
                nome_com_sufixo = f"{nome} - SICOOB"
            else:
                nome_com_sufixo = f"{nome} {cnpj_cpf} - SICOOB"
        elif nome:
            nome_com_sufixo = f"{nome} - SICOOB"
        elif cnpj_cpf:
            nome_com_sufixo = f"{cnpj_cpf} - SICOOB"
        else:
            nome_com_sufixo = "SICOOB"
        
        # Formatar valor no padrão brasileiro (vírgula como separador decimal)
        valor_br = f"{lancamento['valor']:.2f}".replace('.', ',')
        
        dados_csv.append({
            'data': lancamento['data'],
            'nome': nome_com_sufixo,
            'valor': valor_br
        })
    
    # Criar DataFrame e salvar CSV com separador ponto e vírgula
    df_csv = pd.DataFrame(dados_csv)
    df_csv.to_csv(caminho_csv, index=False, encoding='utf-8-sig', sep=';')
    
    print(f"\n📄 CSV GERADO COM SUCESSO!")
    print(f"📁 Arquivo: {caminho_csv}")
    print(f"📈 Total de lançamentos: {len(dados_csv)}")
    
    return str(caminho_csv)

def extrair_texto_e_organizar(caminho_pdf, debug=False):
    """
    Extrai texto do PDF e organiza os lançamentos por data
    
    Args:
        caminho_pdf (str): Caminho para o arquivo PDF
        debug (bool): Se True, exibe informações detalhadas
    """
    # Extrair texto
    texto = extrair_texto_pdf(caminho_pdf, debug)
    
    if texto:
        if debug:
            print("\n" + "=" * 50)
            print("EXTRAÇÃO CONCLUÍDA!")
            print(f"Total de caracteres extraídos: {len(texto)}")
        
        # Organizar lançamentos por data
        lancamentos_por_data = organizar_lancamentos_por_data(texto)
        
        # Exibir lançamentos organizados (apenas em debug)
        exibir_lancamentos_organizados(lancamentos_por_data, debug)
        
        # Processar data e valor
        lancamentos_processados, saldo_anterior, ultimo_saldo_dia = processar_lancamentos_com_data_valor(lancamentos_por_data)
        
        # Exibir lançamentos processados
        exibir_lancamentos_processados(lancamentos_processados, saldo_anterior, ultimo_saldo_dia, debug)
        
        # Gerar planilha Excel
        gerar_planilha_excel(lancamentos_processados, saldo_anterior, ultimo_saldo_dia, caminho_pdf)
        
        # Gerar CSV simplificado
        gerar_csv_simplificado(lancamentos_processados, caminho_pdf)

def gerar_planilha_excel(lancamentos_processados, saldo_anterior, ultimo_saldo_dia, caminho_pdf):
    """
    Gera um arquivo Excel (.xlsx) com os lançamentos processados
    """
    if not lancamentos_processados:
        print("❌ Nenhum lançamento para gerar planilha!")
        return
    
    # Ordenar lançamentos cronologicamente
    lancamentos_ordenados = sorted(lancamentos_processados, key=lambda x: datetime.strptime(x['data'], '%d/%m/%Y'))
    
    # Criar DataFrame
    dados = []
    for lancamento in lancamentos_ordenados:
        dados.append({
            'Data': lancamento['data'],
            'Valor': lancamento['valor'],
            'Tipo': lancamento['tipo'],
            'CNPJ/CPF': lancamento.get('cnpj_cpf', ''),
            'Documento': lancamento.get('documento', ''),
            'Pagador/Recebedor': lancamento.get('pagador_recebedor', ''),
            'Lançamento Completo': lancamento['lancamento_completo']
        })
    
    df = pd.DataFrame(dados)
    
    # Calcular totais
    total_entradas = sum(l['valor'] for l in lancamentos_ordenados if l['valor'] > 0)
    total_saidas = sum(l['valor'] for l in lancamentos_ordenados if l['valor'] < 0)
    saldo_movimentacoes = total_entradas + total_saidas
    
    # Obter diretório e nome do arquivo PDF
    pdf_path = Path(caminho_pdf)
    diretorio = pdf_path.parent
    nome_base = pdf_path.stem
    
    # Nome do arquivo Excel
    nome_excel = f"{nome_base}_EXTRATO_PROCESSADO.xlsx"
    caminho_excel = diretorio / nome_excel
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Extrato Processado"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Adicionar título
    ws['A1'] = f"EXTRATO BANCÁRIO - {nome_base}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    # Adicionar informações de saldo
    ws['A3'] = "RESUMO DE SALDOS"
    ws['A3'].font = Font(bold=True, size=12)
    ws.merge_cells('A3:G3')
    
    if saldo_anterior is not None:
        ws['A4'] = f"Saldo Anterior: R$ {saldo_anterior:,.2f}"
    if ultimo_saldo_dia is not None:
        ws['A5'] = f"Saldo Final: R$ {ultimo_saldo_dia:,.2f}"
    
    ws['A6'] = f"Total Entradas: R$ {total_entradas:,.2f}"
    ws['A7'] = f"Total Saídas: R$ {total_saidas:,.2f}"
    ws['A8'] = f"Saldo Movimentações: R$ {saldo_movimentacoes:,.2f}"
    
    # Adicionar cabeçalhos da tabela
    headers = ['Data', 'Valor', 'Tipo', 'CNPJ/CPF', 'Documento', 'Pagador/Recebedor', 'Lançamento Completo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Adicionar dados
    for row_idx, lancamento in enumerate(lancamentos_ordenados, 11):
        # Data
        ws.cell(row=row_idx, column=1, value=lancamento['data']).border = border
        
        # Valor
        valor_cell = ws.cell(row=row_idx, column=2, value=lancamento['valor'])
        valor_cell.number_format = 'R$ #,##0.00'
        valor_cell.border = border
        
        # Tipo
        ws.cell(row=row_idx, column=3, value=lancamento['tipo']).border = border
        
        # CNPJ/CPF
        ws.cell(row=row_idx, column=4, value=lancamento.get('cnpj_cpf', '')).border = border
        
        # Documento
        ws.cell(row=row_idx, column=5, value=lancamento.get('documento', '')).border = border
        
        # Pagador/Recebedor
        ws.cell(row=row_idx, column=6, value=lancamento.get('pagador_recebedor', '')).border = border
        
        # Lançamento Completo
        ws.cell(row=row_idx, column=7, value=lancamento['lancamento_completo']).border = border
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12  # Data
    ws.column_dimensions['B'].width = 15  # Valor
    ws.column_dimensions['C'].width = 8   # Tipo
    ws.column_dimensions['D'].width = 25  # CNPJ/CPF
    ws.column_dimensions['E'].width = 15  # Documento
    ws.column_dimensions['F'].width = 30  # Pagador/Recebedor
    ws.column_dimensions['G'].width = 80  # Lançamento Completo
    
    # Salvar arquivo
    wb.save(caminho_excel)
    
    print(f"\n📊 PLANILHA GERADA COM SUCESSO!")
    print(f"📁 Arquivo: {caminho_excel}")
    print(f"📈 Total de lançamentos: {len(lancamentos_ordenados)}")
    
    return str(caminho_excel)

def main():
    """Função principal"""
    print("=== EXTRATOR DE TEXTO PDF ===\n")
    
    # Padronização: python conversor_extrato_sicoob_pdf_csv.py <entrada.pdf> <saida.csv>
    if len(sys.argv) < 3:
        print("Uso: python conversor_extrato_sicoob_pdf_csv.py <arquivo.pdf> <arquivo_saida.csv>")
        sys.exit(1)
        caminho_pdf = sys.argv[1]
    caminho_csv = sys.argv[2]
    debug = "--debug" in sys.argv
    if not os.path.exists(caminho_pdf):
        print(f"Erro: O arquivo '{caminho_pdf}' não existe.")
        sys.exit(1)
    if not caminho_pdf.lower().endswith('.pdf'):
        print("Erro: O arquivo deve ser um PDF (.pdf)")
        sys.exit(1)
    texto = extrair_texto_pdf(caminho_pdf, debug)
    if texto:
        lancamentos_por_data = organizar_lancamentos_por_data(texto)
        lancamentos_processados, saldo_anterior, ultimo_saldo_dia = processar_lancamentos_com_data_valor(lancamentos_por_data)
        # Padronizar saída para Data;Histórico;Conta Débito;Conta Crédito;Valor
        def formatar_valor_brl(valor):
            try:
                return f"{float(valor):,.2f}".replace('.', 'X').replace(',', '.').replace('X', ',')
            except Exception:
                return "0,00"
        with open(caminho_csv, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')
            writer.writerow(['Data', 'Histórico', 'Conta Débito', 'Conta Crédito', 'Valor'])
            for l in lancamentos_processados:
                data = l['data']
                valor = l['valor']
                nome = l.get('pagador_recebedor', '')
                if valor > 0:
                    historico = f"RCTO REF {nome.upper()}"
                else:
                    historico = f"PGTO REF {nome.upper()}"
                conta_debito = '1.1.1.01'
                conta_credito = '2.1.1.01'
                writer.writerow([
                    data,
                    historico,
                    conta_debito,
                    conta_credito,
                    formatar_valor_brl(valor)
                ])
        print(f"CSV padronizado gerado em: {caminho_csv}")

if __name__ == "__main__":
    main()
